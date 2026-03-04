"""
Excel Exporter
Export data to Excel files
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
from models.sale import Sale
from models.product import Product


class ExcelExporter:
    """Export data to Excel"""
    
    def __init__(self):
        # Define styles
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        
        self.total_font = Font(bold=True, size=11)
        self.total_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        
        self.border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )
    
    def export_sales(self, sales: list, filename: str = None) -> str:
        """Export sales to Excel"""
        
        if not filename:
            output_dir = Path.home() / '.restaurant_pos' / 'exports'
            output_dir.mkdir(exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = output_dir / f"ventas_{timestamp}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Ventas"
        
        # Headers
        headers = ['Nº Factura', 'Fecha', 'Cliente', 'Método de Pago', 'Subtotal', 'Descuento', 'Total']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Data
        total_subtotal = 0
        total_descuento = 0
        total_total = 0
        
        for row, sale in enumerate(sales, 2):
            fecha = datetime.fromisoformat(sale.fecha_venta).strftime("%d/%m/%Y %H:%M")
            
            ws.cell(row=row, column=1, value=sale.numero_factura).border = self.border
            ws.cell(row=row, column=2, value=fecha).border = self.border
            ws.cell(row=row, column=3, value=sale.cliente).border = self.border
            ws.cell(row=row, column=4, value=sale.metodo_pago.title()).border = self.border
            
            cell = ws.cell(row=row, column=5, value=sale.subtotal)
            cell.number_format = '#,##0.00'
            cell.border = self.border
            
            cell = ws.cell(row=row, column=6, value=sale.descuento)
            cell.number_format = '#,##0.00'
            cell.border = self.border
            
            cell = ws.cell(row=row, column=7, value=sale.total)
            cell.number_format = '#,##0.00'
            cell.border = self.border
            cell.font = Font(bold=True)
            
            total_subtotal += sale.subtotal
            total_descuento += sale.descuento
            total_total += sale.total
        
        # Totals row
        if sales:
            total_row = len(sales) + 2
            
            ws.cell(row=total_row, column=1, value="TOTAL").font = self.total_font
            ws.cell(row=total_row, column=1).fill = self.total_fill
            
            for col in range(2, 5):
                ws.cell(row=total_row, column=col).fill = self.total_fill
            
            cell = ws.cell(row=total_row, column=5, value=total_subtotal)
            cell.number_format = '#,##0.00'
            cell.font = self.total_font
            cell.fill = self.total_fill
            
            cell = ws.cell(row=total_row, column=6, value=total_descuento)
            cell.number_format = '#,##0.00'
            cell.font = self.total_font
            cell.fill = self.total_fill
            
            cell = ws.cell(row=total_row, column=7, value=total_total)
            cell.number_format = '#,##0.00'
            cell.font = self.total_font
            cell.fill = self.total_fill
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        
        # Save
        wb.save(filename)
        print(f"✓ Sales exported to Excel: {filename}")
        return str(filename)
    
    def export_products(self, products: list, filename: str = None) -> str:
        """Export products to Excel"""
        
        if not filename:
            output_dir = Path.home() / '.restaurant_pos' / 'exports'
            output_dir.mkdir(exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = output_dir / f"productos_{timestamp}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos"
        
        # Headers
        headers = ['ID', 'Nombre', 'Descripción', 'Categoría', 'Precio', 'Costo', 'Stock', 'Margen %']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Data
        from models.product import Category
        
        for row, product in enumerate(products, 2):
            category = Category.get_by_id(product.categoria_id)
            cat_name = category.nombre if category else "Sin categoría"
            margen = product.get_margen()
            
            ws.cell(row=row, column=1, value=product.id).border = self.border
            ws.cell(row=row, column=2, value=product.nombre).border = self.border
            ws.cell(row=row, column=3, value=product.descripcion or "").border = self.border
            ws.cell(row=row, column=4, value=cat_name).border = self.border
            
            cell = ws.cell(row=row, column=5, value=product.precio)
            cell.number_format = '#,##0.00'
            cell.border = self.border
            
            cell = ws.cell(row=row, column=6, value=product.costo)
            cell.number_format = '#,##0.00'
            cell.border = self.border
            
            cell = ws.cell(row=row, column=7, value=product.stock)
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center")
            
            # Color code stock
            if product.stock == 0:
                cell.font = Font(color="EF4444", bold=True)
            elif product.stock < 10:
                cell.font = Font(color="F59E0B")
            
            cell = ws.cell(row=row, column=8, value=margen)
            cell.number_format = '0.00"%"'
            cell.border = self.border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 12
        
        # Save
        wb.save(filename)
        print(f"✓ Products exported to Excel: {filename}")
        return str(filename)
    
    def export_sales_summary(self, fecha_desde: str, fecha_hasta: str, filename: str = None) -> str:
        """Export sales summary with statistics"""
        
        if not filename:
            output_dir = Path.home() / '.restaurant_pos' / 'exports'
            output_dir.mkdir(exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = output_dir / f"resumen_ventas_{timestamp}.xlsx"
        
        wb = Workbook()
        
        # Sheet 1: Summary
        ws1 = wb.active
        ws1.title = "Resumen"
        
        summary = Sale.get_sales_summary(fecha_desde, fecha_hasta)
        
        ws1['A1'] = "Reporte de Ventas"
        ws1['A1'].font = Font(bold=True, size=16)
        
        ws1['A3'] = "Periodo:"
        ws1['B3'] = f"{fecha_desde} a {fecha_hasta}"
        
        ws1['A5'] = "Métrica"
        ws1['B5'] = "Valor"
        ws1['A5'].font = self.header_font
        ws1['B5'].font = self.header_font
        ws1['A5'].fill = self.header_fill
        ws1['B5'].fill = self.header_fill
        
        metrics = [
            ("Total de Ventas", summary['total_ventas']),
            ("Ingresos Totales", summary['total_ingresos']),
            ("Descuentos", summary['total_descuentos']),
            ("Venta Promedio", summary['promedio_venta'])
        ]
        
        for i, (metric, value) in enumerate(metrics, 6):
            ws1[f'A{i}'] = metric
            ws1[f'B{i}'] = value
            if i > 6:
                ws1[f'B{i}'].number_format = '#,##0.00'
        
        # Sheet 2: Top Products
        ws2 = wb.create_sheet("Top Productos")
        
        top_products = Sale.get_top_products(20, fecha_desde, fecha_hasta)
        
        ws2['A1'] = "Ranking"
        ws2['B1'] = "Producto"
        ws2['C1'] = "Unidades Vendidas"
        ws2['D1'] = "Ingresos"
        
        for col in range(1, 5):
            cell = ws2.cell(row=1, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
        
        for i, product in enumerate(top_products, 2):
            ws2[f'A{i}'] = i - 1
            ws2[f'B{i}'] = product['nombre']
            ws2[f'C{i}'] = product['total_vendido']
            ws2[f'D{i}'] = product['total_ingresos']
            ws2[f'D{i}'].number_format = '#,##0.00'
        
        ws2.column_dimensions['A'].width = 10
        ws2.column_dimensions['B'].width = 30
        ws2.column_dimensions['C'].width = 18
        ws2.column_dimensions['D'].width = 15
        
        # Save
        wb.save(filename)
        print(f"✓ Sales summary exported to Excel: {filename}")
        return str(filename)


# Convenience functions
def export_sales(sales: list, filename: str = None) -> str:
    """Export sales to Excel"""
    exporter = ExcelExporter()
    return exporter.export_sales(sales, filename)


def export_products(products: list, filename: str = None) -> str:
    """Export products to Excel"""
    exporter = ExcelExporter()
    return exporter.export_products(products, filename)


def export_sales_summary(fecha_desde: str, fecha_hasta: str, filename: str = None) -> str:
    """Export sales summary to Excel"""
    exporter = ExcelExporter()
    return exporter.export_sales_summary(fecha_desde, fecha_hasta, filename)


if __name__ == '__main__':
    # Test Excel export
    from models.product import Product
    
    print("Testing Excel export...")
    
    # Export products
    products = Product.get_all()
    if products:
        excel_file = export_products(products)
        print(f"✓ Test products export: {excel_file}")
