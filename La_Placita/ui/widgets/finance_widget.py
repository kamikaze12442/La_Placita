"""
Finance Widget - Complete Financial Analysis
Monthly revenue, expenses, profit margin, and detailed product analysis
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QFrame,
    QTableWidget, QTableWidgetItem, QPushButton, QComboBox,
    QScrollArea, QHeaderView, QDialog, QDialogButtonBox,
    QLineEdit, QDoubleSpinBox, QMessageBox, QSpinBox,
    QDateEdit, QTextEdit, QFormLayout, QSizePolicy
)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QFont, QColor
from datetime import datetime
from database.connection import db
from utils.excel_exporter import ExcelExporter
from utils.pdf_generator import InvoiceGenerator
import calendar
from pathlib import Path


# ──────────────────────────────────────────────────────────────────────────────
# Smart widgets — evitan scroll accidental al navegar la pantalla
# ──────────────────────────────────────────────────────────────────────────────

class SmartComboBox(QComboBox):
    """ComboBox que solo permite scroll cuando el desplegable está abierto."""
    def wheelEvent(self, event):
        if self.view().isVisible():
            super().wheelEvent(event)
        else:
            event.ignore()


class SmartSpinBox(QSpinBox):
    """SpinBox que solo permite scroll cuando tiene el foco."""
    def wheelEvent(self, event):
        if self.hasFocus():
            super().wheelEvent(event)
        else:
            event.ignore()


class SmartDoubleSpinBox(QDoubleSpinBox):
    """DoubleSpinBox que solo permite scroll cuando tiene el foco."""
    def wheelEvent(self, event):
        if self.hasFocus():
            super().wheelEvent(event)
        else:
            event.ignore()


# ──────────────────────────────────────────────────────────────────────────────
# StatCard
# ──────────────────────────────────────────────────────────────────────────────

class StatCard(QFrame):
    """Financial stat card"""

    def __init__(self, icon, value, label, color="#FF6B35", editable=False, on_edit=None):
        super().__init__()
        self.value_text = value
        self.editable   = editable
        self.on_edit    = on_edit

        self.setObjectName("stat-card")
        self.setStyleSheet(f"""
            QFrame#stat-card {{
                background-color: white;
                border: 1px solid #E5E7EB;
                border-radius: 16px;
                padding: 24px;
                min-height: 120px;
            }}
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(12)

        top_layout = QHBoxLayout()

        icon_label = QLabel(icon)
        icon_label.setStyleSheet(f"""
            background-color: {color}15;
            color: {color};
            font-size: 32px;
            padding: 12px;
            border-radius: 12px;
            min-width: 26px;
            max-width: 26px;
            min-height: 26px;
            max-height: 26px;
        """)
        icon_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        top_layout.addWidget(icon_label)

        if editable:
            edit_btn = QPushButton("✏️")
            edit_btn.setFixedSize(26, 10)
            edit_btn.setStyleSheet("""
                QPushButton {
                    background-color: #3B82F6;
                    color: white;
                    border-radius: 6px;
                    min-width: 26px; max-width: 26px;
                    min-height: 26px; max-height: 26px;
                    margin: 0px 0px 0px 80px;
                }
                QPushButton:hover { background-color: #2563EB; }
            """)
            edit_btn.clicked.connect(self.on_edit)
            top_layout.addWidget(edit_btn)

        top_layout.addStretch()
        layout.addLayout(top_layout)

        self.value_label = QLabel(value)
        self.value_label.setStyleSheet(
            "font-size: 28px; font-weight: 700; color: #1F2937;")
        layout.addWidget(self.value_label)

        label_label = QLabel(label)
        label_label.setStyleSheet(
            "font-size: 14px; color: black; font-weight: 500;")
        layout.addWidget(label_label)

        layout.addStretch()

    def update_value(self, value):
        self.value_text = value
        self.value_label.setText(value)


# ──────────────────────────────────────────────────────────────────────────────
# Constantes compartidas
# ──────────────────────────────────────────────────────────────────────────────

CATEGORIAS_COSTO = [
    ("🏠", "Alquiler / Infraestructura"),
    ("👥", "Personal / Sueldos"),
    ("💡", "Servicios (agua, luz, gas)"),
    ("📦", "Otros"),
]

EJEMPLOS_COSTO = {
    "fijo":     "• Alquiler&nbsp;&nbsp;• Sueldos fijos&nbsp;&nbsp;• Seguros&nbsp;&nbsp;• Servicios básicos",
    "variable": "• Insumos extra&nbsp;&nbsp;• Horas extra&nbsp;&nbsp;• Comisiones&nbsp;&nbsp;• Gastos eventuales",
}

MESES_ES = [
    'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
]


# ──────────────────────────────────────────────────────────────────────────────
# AddCostDialog — dialog para agregar un costo individual
# ──────────────────────────────────────────────────────────────────────────────

class AddCostDialog(QDialog):
    """Dialog para agregar un nuevo costo (fijo o variable)"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Agregar Costo")
        self.setMinimumWidth(480)
        self.setModal(True)
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(18)
        layout.setContentsMargins(28, 24, 28, 24)

        title = QLabel("💼 Nuevo Costo")
        title.setStyleSheet("font-size: 20px; font-weight: 700; color: #1F2937;")
        layout.addWidget(title)

        form = QFormLayout()
        form.setSpacing(14)
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)

        # Tipo
        self.tipo_combo = SmartComboBox()
        self.tipo_combo.addItem("🏢  Costo Fijo",     "fijo")
        self.tipo_combo.addItem("📊  Costo Variable", "variable")
        self.tipo_combo.setStyleSheet("font-size: 14px; padding: 6px;")
        self.tipo_combo.currentIndexChanged.connect(self._update_info)
        form.addRow("Tipo: *", self.tipo_combo)

        # Concepto
        self.concepto_input = QLineEdit()
        self.concepto_input.setPlaceholderText("Ej: Alquiler local, Sueldo mesero…")
        self.concepto_input.setStyleSheet("font-size: 14px; padding: 6px;")
        form.addRow("Concepto: *", self.concepto_input)

        # Categoría
        self.categoria_combo = SmartComboBox()
        for icono, nombre in CATEGORIAS_COSTO:
            self.categoria_combo.addItem(f"{icono}  {nombre}", nombre)
        self.categoria_combo.setStyleSheet("font-size: 14px; padding: 6px;")
        form.addRow("Categoría:", self.categoria_combo)

        # Monto
        self.monto_input = SmartDoubleSpinBox()
        self.monto_input.setPrefix("Bs ")
        self.monto_input.setRange(0.01, 999999.99)
        self.monto_input.setDecimals(2)
        self.monto_input.setSingleStep(10)
        self.monto_input.setStyleSheet("font-size: 15px; padding: 6px;")
        form.addRow("Monto: *", self.monto_input)

        # Fecha
        self.fecha_input = QDateEdit()
        self.fecha_input.setCalendarPopup(True)
        self.fecha_input.setDate(QDate.currentDate())
        self.fecha_input.setDisplayFormat("dd/MM/yyyy")
        self.fecha_input.setStyleSheet("font-size: 14px; padding: 6px;")
        form.addRow("Fecha:", self.fecha_input)

        # Descripción
        self.desc_input = QTextEdit()
        self.desc_input.setMaximumHeight(70)
        self.desc_input.setPlaceholderText("Nota adicional (opcional)…")
        self.desc_input.setStyleSheet("font-size: 13px; padding: 4px;")
        form.addRow("Descripción:", self.desc_input)

        layout.addLayout(form)

        # Panel informativo contextual
        self.info_label = QLabel()
        self.info_label.setStyleSheet("""
            color: #6B7280; font-size: 13px; padding: 12px;
            background-color: #F9FAFB; border-radius: 8px;
            border: 1px solid #E5E7EB;
        """)
        self.info_label.setWordWrap(True)
        self.info_label.setTextFormat(Qt.TextFormat.RichText)
        layout.addWidget(self.info_label)
        self._update_info()

        btn_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save |
            QDialogButtonBox.StandardButton.Cancel
        )
        btn_box.button(QDialogButtonBox.StandardButton.Save).setText("💾 Guardar")
        btn_box.button(QDialogButtonBox.StandardButton.Save).setDefault(True)
        btn_box.button(QDialogButtonBox.StandardButton.Cancel).setText("Cancelar")
        btn_box.accepted.connect(self._validate_and_accept)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

    def _update_info(self):
        tipo  = self.tipo_combo.currentData()
        color = "#3B82F6" if tipo == "fijo" else "#F59E0B"
        desc  = ("No varía con el volumen de ventas."
                 if tipo == "fijo" else
                 "Cambia según la actividad del negocio.")
        self.info_label.setText(
            f"<b style='color:{color}'>Costo {'fijo' if tipo=='fijo' else 'variable'}:</b> {desc}"
            f"<br><span style='color:#9CA3AF'>Ejemplos: {EJEMPLOS_COSTO.get(tipo,'')}</span>"
        )

    def _validate_and_accept(self):
        if not self.concepto_input.text().strip():
            QMessageBox.warning(self, "Campo requerido", "El concepto es obligatorio.")
            self.concepto_input.setFocus()
            return
        if self.monto_input.value() <= 0:
            QMessageBox.warning(self, "Campo requerido", "El monto debe ser mayor a 0.")
            self.monto_input.setFocus()
            return
        self.accept()

    def get_data(self) -> dict:
        return {
            "tipo":        self.tipo_combo.currentData(),
            "concepto":    self.concepto_input.text().strip(),
            "categoria":   self.categoria_combo.currentData(),
            "monto":       self.monto_input.value(),
            "fecha_gasto": self.fecha_input.date().toString("yyyy-MM-dd"),
            "descripcion": self.desc_input.toPlainText().strip(),
        }


# ──────────────────────────────────────────────────────────────────────────────
# VerCostosDialog — dialog para ver, filtrar y eliminar costos registrados
# ──────────────────────────────────────────────────────────────────────────────

class VerCostosDialog(QDialog):
    """Dialog con tabla filtrable de todos los costos del período seleccionado."""

    def __init__(self, current_year: int, current_month: int, parent=None):
        super().__init__(parent)
        self.setWindowTitle("📋 Costos Registrados")
        self.setMinimumSize(780, 560)
        self.setModal(True)
        self._year          = current_year
        self._month         = current_month
        self._all_rows      = []
        self._finance_widget = parent   # referencia directa, no depende de self.parent()
        self._init_ui()
        self._load_data()

    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(16)
        layout.setContentsMargins(24, 20, 24, 20)

        # Título
        title = QLabel("📋 Costos Registrados")
        title.setStyleSheet("font-size: 20px; font-weight: 700; color: #1F2937;")
        layout.addWidget(title)

        # ── Barra de filtros ──────────────────────────────────────────
        filter_frame = QFrame()
        filter_frame.setStyleSheet("""
            QFrame {
                background-color: #F9FAFB;
                border: 1px solid #E5E7EB;
                border-radius: 10px;
            }
        """)
        fl = QHBoxLayout(filter_frame)
        fl.setSpacing(10)
        fl.setContentsMargins(14, 10, 14, 10)

        def lbl(text):
            l = QLabel(text)
            l.setStyleSheet("font-weight: 600; color: #4B5563; font-size: 13px;")
            return l

        # Mes
        fl.addWidget(lbl("Mes:"))
        self.f_month = SmartComboBox()
        self.f_month.addItem("Todos", 0)
        for i, m in enumerate(MESES_ES, 1):
            self.f_month.addItem(m, i)
        self.f_month.setCurrentIndex(self._month)   # preseleccionar mes actual
        self.f_month.currentIndexChanged.connect(self._apply_filters)
        fl.addWidget(self.f_month)

        # Año
        fl.addWidget(lbl("Año:"))
        self.f_year = SmartSpinBox()
        self.f_year.setRange(2020, 2030)
        self.f_year.setValue(self._year)
        self.f_year.valueChanged.connect(self._on_year_changed)
        fl.addWidget(self.f_year)

        # Tipo
        fl.addWidget(lbl("Tipo:"))
        self.f_tipo = SmartComboBox()
        self.f_tipo.addItem("Todos",       "todos")
        self.f_tipo.addItem("🏢 Fijo",     "fijo")
        self.f_tipo.addItem("📊 Variable", "variable")
        self.f_tipo.currentIndexChanged.connect(self._apply_filters)
        fl.addWidget(self.f_tipo)

        # Categoría
        fl.addWidget(lbl("Categoría:"))
        self.f_cat = SmartComboBox()
        self.f_cat.addItem("Todas", "todas")
        for _, nombre in CATEGORIAS_COSTO:
            self.f_cat.addItem(nombre, nombre)
        self.f_cat.currentIndexChanged.connect(self._apply_filters)
        fl.addWidget(self.f_cat)

        fl.addStretch()

        clear_btn = QPushButton("↺ Limpiar")
        clear_btn.setStyleSheet("""
            QPushButton {
                background-color: #E5E7EB; color: #4B5563;
                padding: 6px 14px; border-radius: 7px;
                font-weight: 600; font-size: 12px;
            }
            QPushButton:hover { background-color: #D1D5DB; }
        """)
        clear_btn.clicked.connect(self._clear_filters)
        fl.addWidget(clear_btn)

        layout.addWidget(filter_frame)

        # Resumen
        self.resumen_lbl = QLabel()
        self.resumen_lbl.setStyleSheet("font-size: 13px; color: #6B7280;")
        self.resumen_lbl.setTextFormat(Qt.TextFormat.RichText)
        layout.addWidget(self.resumen_lbl)

        # ── Tabla ─────────────────────────────────────────────────────
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(
            ["Tipo", "Concepto", "Categoría", "Monto", "Fecha", ""]
        )
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setShowGrid(False)
        self.table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #E5E7EB; border-radius: 10px;
                font-size: 13px; background-color: white;
                alternate-background-color: #F9FAFB;
            }
            QTableWidget::item { padding: 8px; border-bottom: 1px solid #F3F4F6; }
            QHeaderView::section {
                background-color: #F3F4F6; color: #6B7280;
                font-weight: 600; font-size: 12px;
                padding: 10px 8px; border: none;
                border-bottom: 2px solid #E5E7EB;
            }
        """)

        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        hdr.setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(5, 48)

        layout.addWidget(self.table)

        # Botón cerrar
        close_btn = QPushButton("Cerrar")
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #6B7280; color: white;
                padding: 10px 28px; border-radius: 8px;
                font-weight: 600; font-size: 13px;
            }
            QPushButton:hover { background-color: #4B5563; }
        """)
        close_btn.clicked.connect(self.accept)
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)

    # ── Carga y filtrado ──────────────────────────────────────────────

    def _load_data(self):
        """Carga todos los costos del año desde la BD."""
        year  = self.f_year.value()
        rows  = db.fetch_all(
            "SELECT id, tipo, concepto, categoria, monto, fecha_gasto "
            "FROM gastos WHERE strftime('%Y', fecha_gasto)=? "
            "ORDER BY fecha_gasto DESC, id DESC",
            (str(year),)
        )
        self._all_rows = [dict(r) for r in rows] if rows else []
        self._apply_filters()

    def _on_year_changed(self):
        """Al cambiar el año vuelve a consultar la BD."""
        self._load_data()

    def _apply_filters(self):
        mes_sel  = self.f_month.currentData()   # 0 = todos
        tipo_sel = self.f_tipo.currentData()
        cat_sel  = self.f_cat.currentData()

        filtered = []
        for r in self._all_rows:
            if mes_sel != 0:
                if int(r["fecha_gasto"][5:7]) != mes_sel:
                    continue
            if tipo_sel != "todos" and r["tipo"] != tipo_sel:
                continue
            if cat_sel != "todas" and r.get("categoria", "") != cat_sel:
                continue
            filtered.append(r)

        self._fill_table(filtered)

    def _fill_table(self, rows: list):
        self.table.setRowCount(len(rows))
        total_fijo = total_var = 0.0

        for i, r in enumerate(rows):
            es_fijo = r["tipo"] == "fijo"
            monto   = float(r["monto"])
            if es_fijo:
                total_fijo += monto
            else:
                total_var  += monto

            # Tipo badge
            tipo_lbl = QLabel("🏢 Fijo" if es_fijo else "📊 Variable")
            tipo_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            tipo_lbl.setStyleSheet(
                f"background-color:{'#EFF6FF' if es_fijo else '#FFFBEB'};"
                f"color:{'#3B82F6' if es_fijo else '#D97706'};"
                "font-size:12px;font-weight:600;"
                "padding:4px 8px;border-radius:6px;margin:4px;"
            )
            self.table.setCellWidget(i, 0, tipo_lbl)

            self.table.setItem(i, 1, QTableWidgetItem(r["concepto"]))

            cat_item = QTableWidgetItem(r.get("categoria") or "—")
            cat_item.setForeground(QColor("#6B7280"))
            self.table.setItem(i, 2, cat_item)

            monto_item = QTableWidgetItem(f"Bs {monto:.2f}")
            monto_item.setTextAlignment(
                Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            monto_item.setFont(QFont("Segoe UI", 13, QFont.Weight.Bold))
            self.table.setItem(i, 3, monto_item)

            try:
                fecha_fmt = datetime.strptime(
                    r["fecha_gasto"], "%Y-%m-%d").strftime("%d/%m/%Y")
            except Exception:
                fecha_fmt = r["fecha_gasto"]
            fecha_item = QTableWidgetItem(fecha_fmt)
            fecha_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            fecha_item.setForeground(QColor("#6B7280"))
            self.table.setItem(i, 4, fecha_item)

            del_btn = QPushButton("🗑️")
            del_btn.setToolTip("Eliminar")
            del_btn.setStyleSheet("""
                QPushButton {
                    background-color:#FEE2E2;color:#EF4444;
                    border:none;border-radius:6px;
                    font-size:14px;padding:4px;margin:4px;
                }
                QPushButton:hover{background-color:#FECACA;}
            """)
            gid = r["id"]
            del_btn.clicked.connect(lambda checked=False, g=gid: self._delete_cost(g))
            self.table.setCellWidget(i, 5, del_btn)
            self.table.setRowHeight(i, 44)

        total = total_fijo + total_var
        self.resumen_lbl.setText(
            f"{len(rows)} registro(s)  ·  "
            f"Fijos: <b>Bs {total_fijo:.2f}</b>  |  "
            f"Variables: <b>Bs {total_var:.2f}</b>  |  "
            f"Total: <b>Bs {total:.2f}</b>"
        )

    def _delete_cost(self, gasto_id: int):
        reply = QMessageBox.question(
            self, "Confirmar eliminación",
            "¿Eliminar este costo? Esta acción no se puede deshacer.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            try:
                gasto_id = int(gasto_id)
                conn = db.get_connection()
                conn.execute("DELETE FROM gastos WHERE id = ?", (gasto_id,))
                conn.commit()
                self._all_rows = [r for r in self._all_rows if int(r["id"]) != gasto_id]
                self._apply_filters()
                if self._finance_widget and hasattr(self._finance_widget, "refresh_after_cost_change"):
                    self._finance_widget.refresh_after_cost_change()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"No se pudo eliminar el costo:\n{str(e)}")

    def _clear_filters(self):
        self.f_month.setCurrentIndex(0)
        self.f_tipo.setCurrentIndex(0)
        self.f_cat.setCurrentIndex(0)


# ──────────────────────────────────────────────────────────────────────────────
# CostosCard — card compacta (misma fila que Ingresos / Gastos / Ganancia)
# ──────────────────────────────────────────────────────────────────────────────

class CostosCard(QFrame):
    """Card compacta de Costos con totales y botones Agregar / Ver."""

    def __init__(self, parent_widget):
        super().__init__()
        self._parent         = parent_widget
        self._year           = datetime.now().year
        self._month          = datetime.now().month
        self._total_fijo     = 0.0
        self._total_variable = 0.0

        self.setObjectName("costos-stat-card")
        self.setStyleSheet("""
            QFrame#costos-stat-card {
                background-color: white;
                border: 1px solid #E5E7EB;
                border-radius: 16px;
            }
        """)
        self.setFixedSize(300, 250)
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(8)

        # ── Ícono fijo 52x52 ──────────────────────────────────────────
        icon_lbl = QLabel("💼")
        icon_lbl.setFixedSize(52, 52)
        icon_lbl.setStyleSheet("""
            background-color: #8B5CF620;
            color: #8B5CF6;
            font-size: 24px;
            border-radius: 12px;
        """)
        icon_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(icon_lbl)

        # ── Valor total ───────────────────────────────────────────────
        self.total_lbl = QLabel("Bs 0.00")
        self.total_lbl.setFixedHeight(38)
        self.total_lbl.setStyleSheet(
            "font-size: 26px; font-weight: 700; color: #1F2937;")
        layout.addWidget(self.total_lbl)

        # ── Título ────────────────────────────────────────────────────
        title_lbl = QLabel("Costos del Mes")
        title_lbl.setFixedHeight(20)
        title_lbl.setStyleSheet(
            "font-size: 13px; color: #111827; font-weight: 500;")
        layout.addWidget(title_lbl)

        # ── Desglose Fijo / Variable ──────────────────────────────────
        self.detalle_lbl = QLabel("Fijo: Bs 0.00  |  Variable: Bs 0.00")
        self.detalle_lbl.setFixedHeight(18)
        self.detalle_lbl.setStyleSheet("font-size: 11px; color: #6B7280;")
        layout.addWidget(self.detalle_lbl)

        layout.addStretch()

        # ── Botones al pie ────────────────────────────────────────────
        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)
        btn_row.setContentsMargins(0, 0, 0, 0)

        ver_btn = QPushButton("📋 Ver Costos")
        ver_btn.setFixedHeight(32)
        ver_btn.setToolTip("Ver y filtrar costos registrados")
        ver_btn.setStyleSheet("""
            QPushButton {
                background-color: #E0E7FF; color: #4F46E5;
                padding: 0px 12px; border-radius: 8px;
                font-weight: 600; font-size: 12px;
            }
            QPushButton:hover { background-color: #C7D2FE; }
        """)
        ver_btn.clicked.connect(self._open_ver_dialog)
        btn_row.addWidget(ver_btn)

        add_btn = QPushButton("➕ Agregar")
        add_btn.setFixedHeight(32)
        add_btn.setToolTip("Registrar un nuevo costo")
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #8B5CF6; color: white;
                padding: 0px 12px; border-radius: 8px;
                font-weight: 600; font-size: 12px;
            }
            QPushButton:hover { background-color: #7C3AED; }
        """)
        add_btn.clicked.connect(self._open_add_dialog)
        btn_row.addWidget(add_btn)

        layout.addLayout(btn_row)

    # ── Carga ─────────────────────────────────────────────────────────

    def load_costs(self, year: int, month: int):
        self._year  = year
        self._month = month
        rows = db.fetch_all(
            "SELECT tipo, COALESCE(SUM(monto),0) as total "
            "FROM gastos WHERE strftime('%Y-%m', fecha_gasto)=? GROUP BY tipo",
            (f"{year}-{month:02d}",)
        )
        self._total_fijo     = 0.0
        self._total_variable = 0.0
        for r in (rows or []):
            if r["tipo"] == "fijo":
                self._total_fijo = float(r["total"])
            elif r["tipo"] == "variable":
                self._total_variable = float(r["total"])
        self._refresh_display()

    def _refresh_display(self):
        total = self._total_fijo + self._total_variable
        self.total_lbl.setText(f"Bs {total:.2f}")
        self.detalle_lbl.setText(
            f"Fijo: Bs {self._total_fijo:.2f}  |  Variable: Bs {self._total_variable:.2f}"
        )

    def get_totals(self) -> tuple:
        return self._total_fijo, self._total_variable

    # ── Acciones ──────────────────────────────────────────────────────

    def _open_add_dialog(self):
        dialog = AddCostDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            data = dialog.get_data()
            # ✅ execute_query es el método correcto del DatabaseManager
            db.execute_query(
                "INSERT INTO gastos "
                "(concepto, monto, tipo, categoria, descripcion, fecha_gasto, usuario_id) "
                "VALUES (?, ?, ?, ?, ?, ?, ?)",
                (
                    data["concepto"],
                    data["monto"],
                    data["tipo"],
                    data["categoria"],
                    data["descripcion"],
                    data["fecha_gasto"],
                    1,
                ),
            )
            self.load_costs(self._year, self._month)
            self._parent.refresh_after_cost_change()

    def _open_ver_dialog(self):
        dialog = VerCostosDialog(self._year, self._month, parent=self._parent)
        dialog.exec()
        # Refrescar por si hubo eliminaciones desde el dialog
        self.load_costs(self._year, self._month)
        self._parent.refresh_after_cost_change()


# ──────────────────────────────────────────────────────────────────────────────
# FinanceWidget
# ──────────────────────────────────────────────────────────────────────────────

class FinanceWidget(QWidget):
    """Complete financial analysis widget"""

    def __init__(self):
        super().__init__()
        self.current_month = datetime.now().month
        self.current_year  = datetime.now().year
        self.init_ui()
        self.update_stats()

    def init_ui(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)

        content = QWidget()
        main_layout = QVBoxLayout(content)
        main_layout.setContentsMargins(40, 30, 40, 40)
        main_layout.setSpacing(30)

        self._create_header(main_layout)

        # Fila 1: 4 stat cards dinámicas (Ingresos / Gastos / Ganancia / Margen)
        self.main_stats_layout = QHBoxLayout()
        self.main_stats_layout.setSpacing(20)
        main_layout.addLayout(self.main_stats_layout)

        # Fila 2: CostosCard alineada bajo "Ingresos del Mes" y "Gastos del Mes"
        self.costos_row_layout = QHBoxLayout()
        self.costos_row_layout.setSpacing(20)
        self.costos_card = CostosCard(parent_widget=self)
        self.costos_card.load_costs(self.current_year, self.current_month)
        self.costos_row_layout.addWidget(self.costos_card)
        self.costos_row_layout.addStretch()
        main_layout.addLayout(self.costos_row_layout)

        self._create_analysis_section(main_layout)

        scroll.setWidget(content)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(scroll)

    # ── Header ────────────────────────────────────────────────────────

    def _create_header(self, layout):
        vl = QVBoxLayout()
        vl.setSpacing(8)
        title = QLabel("Finanzas")
        title.setStyleSheet("font-size: 32px; font-weight: 700; color: #1F2937;")
        vl.addWidget(title)
        sub = QLabel("Análisis financiero completo y control de gastos")
        sub.setStyleSheet("font-size: 14px; color: #6B7280;")
        vl.addWidget(sub)
        layout.addLayout(vl)

    # ── Sección análisis ──────────────────────────────────────────────

    def _create_analysis_section(self, layout):
        frame = QFrame()
        frame.setStyleSheet("""
            QFrame {
                background-color: white;
                border: 1px solid #E5E7EB;
                border-radius: 16px;
                padding: 24px;
            }
        """)
        al = QVBoxLayout(frame)
        al.setSpacing(20)

        # Header con controles
        hl = QHBoxLayout()

        title = QLabel("📊 Análisis Financiero")
        title.setStyleSheet("font-size: 20px; font-weight: 700; color: #1F2937; border: none")
        hl.addWidget(title)
        hl.addStretch()

        hl.addWidget(self._lbl("Mes:"))
        self.month_combo = SmartComboBox()
        for i, m in enumerate(MESES_ES, 1):
            self.month_combo.addItem(m, i)
        self.month_combo.setCurrentIndex(self.current_month - 1)
        self.month_combo.currentIndexChanged.connect(self.on_month_changed)
        hl.addWidget(self.month_combo)

        hl.addWidget(self._lbl("Año:"))
        self.year_spin = SmartSpinBox()
        self.year_spin.setRange(2020, 2030)
        self.year_spin.setValue(self.current_year)
        self.year_spin.valueChanged.connect(self.on_year_changed)
        hl.addWidget(self.year_spin)

        excel_btn = QPushButton("📊 Excel")
        excel_btn.setStyleSheet("""
            QPushButton {
                background-color:#10B981;color:white;
                padding:10px 20px;border-radius:8px;font-weight:600;
            }
            QPushButton:hover{background-color:#059669;}
        """)
        excel_btn.clicked.connect(self.export_excel)
        hl.addWidget(excel_btn)

        pdf_btn = QPushButton("📄 PDF")
        pdf_btn.setStyleSheet("""
            QPushButton {
                background-color:#EF4444;color:white;
                padding:10px 20px;border-radius:8px;font-weight:600;
            }
            QPushButton:hover{background-color:#DC2626;}
        """)
        pdf_btn.clicked.connect(self.export_pdf)
        hl.addWidget(pdf_btn)

        al.addLayout(hl)

        # Filtros: categoría + día
        fl = QHBoxLayout()
        fl.setSpacing(12)

        fl.addWidget(self._lbl("Categoría:"))
        self.category_combo = SmartComboBox()
        self.category_combo.addItem("Todas las categorías", None)
        self._load_categories()
        self.category_combo.currentIndexChanged.connect(self.on_category_changed)
        fl.addWidget(self.category_combo)

        fl.addWidget(self._lbl("Día:"))
        self.day_combo = SmartComboBox()
        self.day_combo.addItem("Todos los días", 0)
        for d in range(1, 32):
            self.day_combo.addItem(str(d), d)
        self.day_combo.currentIndexChanged.connect(self.on_day_changed)
        fl.addWidget(self.day_combo)

        fl.addStretch()
        al.addLayout(fl)

        # Tabla
        self.analysis_table = QTableWidget()
        self.analysis_table.setColumnCount(5)
        self.analysis_table.setHorizontalHeaderLabels(
            ["Mes/Producto", "Ingresos (Bs)", "Gastos (Bs)",
             "Ganancia Neta (Bs)", "Margen (%)"]
        )
        self.analysis_table.setEditTriggers(
            QTableWidget.EditTrigger.NoEditTriggers)
        self.analysis_table.setAlternatingRowColors(True)
        self.analysis_table.setMinimumHeight(500)
        self.analysis_table.setStyleSheet("""
            QTableWidget {
                border: none; gridline-color: #E5E7EB;
                font-size: 13px; background-color: white;
            }
            QTableWidget::item {
                padding: 10px; border-bottom: 1px solid #E5E7EB;
            }
            QTableWidget::item:alternate { background-color: #F9FAFB; }
        """)

        hdr = self.analysis_table.horizontalHeader()
        hdr.setStyleSheet("""
            QHeaderView {
                background-color: white; color: white;
                font-weight: bold; font-size: 20px;
                padding: 10px; border: none;
            }
        """)
        hdr.setMinimumHeight(65)
        hdr.setMaximumHeight(65)
        hdr.setDefaultAlignment(
            Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
        hdr.setStretchLastSection(True)
        for col in range(5):
            hdr.setSectionResizeMode(col, QHeaderView.ResizeMode.ResizeToContents)
        self.analysis_table.setColumnWidth(1, 160)
        self.analysis_table.setColumnWidth(2, 150)
        self.analysis_table.setColumnWidth(3, 190)
        self.analysis_table.setColumnWidth(4, 130)
        self.analysis_table.setMinimumHeight(800)

        al.addWidget(self.analysis_table)
        layout.addWidget(frame)

    def _lbl(self, text):
        l = QLabel(text)
        l.setStyleSheet("font-weight: bold; color: #4B5563;border: none")
        return l

    def _load_categories(self):
        cats = db.fetch_all(
            "SELECT id, nombre FROM categorias WHERE activo=1 ORDER BY nombre")
        if cats:
            for cat in cats:
                self.category_combo.addItem(cat['nombre'], cat['id'])

    # ── Stats principales ─────────────────────────────────────────────

    def update_stats(self):
        """Reconstruye las 4 stat cards de la fila 1."""
        while self.main_stats_layout.count():
            item = self.main_stats_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        month_str = f"{self.current_year}-{self.current_month:02d}"

        inc = db.fetch_one(
            "SELECT COALESCE(SUM(total),0) as total FROM ventas "
            "WHERE strftime('%Y-%m',fecha_venta)=? AND estado='completada'",
            (month_str,)
        )
        ingresos = float(inc['total']) if inc else 0.0

        cst = db.fetch_one(
            "SELECT COALESCE(SUM(dv.cantidad*p.costo),0) as total_costo "
            "FROM detalle_ventas dv "
            "JOIN productos p ON dv.producto_id=p.id "
            "JOIN ventas v    ON dv.venta_id=v.id "
            "WHERE strftime('%Y-%m',v.fecha_venta)=? AND v.estado='completada'",
            (month_str,)
        )
        gastos_prod = float(cst['total_costo']) if cst else 0.0

        costo_fijo, costo_variable = self.costos_card.get_totals()
        gastos_total = gastos_prod + costo_fijo + costo_variable

        ganancia = ingresos - gastos_total
        margen   = (ganancia / ingresos * 100) if ingresos > 0 else 0.0

        cards = [
            StatCard("💰", f"Bs {ingresos:.2f}",     "Ingresos del Mes",  "#10B981"),
            StatCard("💸", f"Bs {gastos_total:.2f}", "Gastos del Mes",    "#EF4444"),
            StatCard("📈", f"Bs {ganancia:.2f}",      "Ganancia Neta",
                     "#10B981" if ganancia >= 0 else "#EF4444"),
            StatCard("📊", f"{margen:.1f}%",           "Margen de Ganancia",
                     "#10B981" if margen >= 0 else "#EF4444"),
        ]
        for card in cards:
            self.main_stats_layout.addWidget(card)

        self.update_analysis_table()

    # ── Tabla de análisis ─────────────────────────────────────────────
    def _set_category_row(self, row, label):
        label_item = QTableWidgetItem(label)
        label_item.setFont(QFont("Segoe UI", 14, QFont.Weight.Bold))
        label_item.setForeground(QColor("#1F2937"))
        label_item.setBackground(QColor("#F3F4F6"))
        self.analysis_table.setItem(row, 0, label_item)
        for col in range(1, 5):
            empty_item = QTableWidgetItem("")
            empty_item.setBackground(QColor("#F3F4F6"))
            self.analysis_table.setItem(row, col, empty_item)
        self.analysis_table.setRowHeight(row, 45)
    def _set_table_row(self, row, label, ingresos, gastos, ganancia, margen, unidades, is_summary=False):
        """Set table row data with optional units sold"""
        # Column 0: Label (Mes/Producto) with units
        if unidades is not None:
            label_text = f"{label}  ({unidades} unidades)"
        else:
            label_text = label
        
        label_item = QTableWidgetItem(label_text)
        
        if is_summary:
            # Summary row (header row for month)
            font = QFont("Segoe UI", 16, QFont.Weight.Bold)
            label_item.setFont(font)
            label_item.setBackground(QColor("#F9FAFB"))
            label_item.setForeground(QColor("#1F2937"))
        else:
            # Product rows
            if "📁" in label:
                # Category header
                font = QFont("Segoe UI", 13, QFont.Weight.Bold)
                label_item.setFont(font)
                label_item.setForeground(QColor("#1F2937"))
                label_item.setBackground(QColor("#F3F4F6"))
            else:
                # Regular product
                font = QFont("Segoe UI", 13)
                label_item.setFont(font)
                label_item.setForeground(QColor("#4B5563"))
        
        self.analysis_table.setItem(row, 0, label_item)
        
        # Column 1: Ingresos
        ing_item = QTableWidgetItem(f"Bs {ingresos:,.2f}")
        ing_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        ing_item.setForeground(QColor("#10B981"))
        
        if is_summary:
            ing_item.setFont(QFont("Segoe UI", 15, QFont.Weight.Bold))
            ing_item.setBackground(QColor("#ECFDF5"))
        elif "📁" in label:
            ing_item.setFont(QFont("Segoe UI", 13, QFont.Weight.Bold))
            ing_item.setBackground(QColor("#F3F4F6"))
        else:
            ing_item.setFont(QFont("Segoe UI", 12))
        
        self.analysis_table.setItem(row, 1, ing_item)
        
        # Column 2: Gastos
        gas_item = QTableWidgetItem(f"Bs {gastos:,.2f}")
        gas_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        gas_item.setForeground(QColor("#EF4444"))
        
        if is_summary:
            gas_item.setFont(QFont("Segoe UI", 15, QFont.Weight.Bold))
            gas_item.setBackground(QColor("#FEE2E2"))
        elif "📁" in label:
            gas_item.setFont(QFont("Segoe UI", 13, QFont.Weight.Bold))
            gas_item.setBackground(QColor("#F3F4F6"))
        else:
            gas_item.setFont(QFont("Segoe UI", 12))
        
        self.analysis_table.setItem(row, 2, gas_item)
        
        # Column 3: Ganancia Neta
        gan_item = QTableWidgetItem(f"Bs {ganancia:,.2f}")
        gan_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        gan_item.setForeground(QColor("#10B981" if ganancia >= 0 else "#EF4444"))
        
        if is_summary:
            gan_item.setFont(QFont("Segoe UI", 15, QFont.Weight.Bold))
            color_bg = "#ECFDF5" if ganancia >= 0 else "#FEE2E2"
            gan_item.setBackground(QColor(color_bg))
        elif "📁" in label:
            gan_item.setFont(QFont("Segoe UI", 13, QFont.Weight.Bold))
            gan_item.setBackground(QColor("#F3F4F6"))
        else:
            gan_item.setFont(QFont("Segoe UI", 12))
        
        self.analysis_table.setItem(row, 3, gan_item)
        
        # Column 4: Margen
        mar_item = QTableWidgetItem(f"{margen:.1f}%")
        mar_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        mar_item.setForeground(QColor("#10B981" if margen >= 0 else "#EF4444"))
        
        if is_summary:
            mar_item.setFont(QFont("Segoe UI", 15, QFont.Weight.Bold))
            color_bg = "#ECFDF5" if margen >= 0 else "#FEE2E2"
            mar_item.setBackground(QColor(color_bg))
        elif "📁" in label:
            mar_item.setFont(QFont("Segoe UI", 13, QFont.Weight.Bold))
            mar_item.setBackground(QColor("#F3F4F6"))
        else:
            mar_item.setFont(QFont("Segoe UI", 12))
        
        self.analysis_table.setItem(row, 4, mar_item)
        
        # Set row height
        if is_summary:
            self.analysis_table.setRowHeight(row, 75)
        elif "📁" in label:
            self.analysis_table.setRowHeight(row, 55)
        else:
            self.analysis_table.setRowHeight(row, 48)
    def update_analysis_table(self):
        month_str = f"{self.current_year}-{self.current_month:02d}"
        selected_day = self.day_combo.currentData()

        if selected_day and selected_day > 0:
            date_filter_v = f"strftime('%Y-%m-%d', v.fecha_venta)='{self.current_year}-{self.current_month:02d}-{selected_day:02d}'"
            date_filter   = f"strftime('%Y-%m-%d', fecha_venta)='{self.current_year}-{self.current_month:02d}-{selected_day:02d}'"
            label_periodo = f"📅 {selected_day:02d}/{self.current_month:02d}/{self.current_year}"
        else:
            date_filter   = f"strftime('%Y-%m', fecha_venta)='{month_str}'"
            date_filter_v = f"strftime('%Y-%m', v.fecha_venta)='{month_str}'"
            label_periodo = f"📅 {MESES_ES[self.current_month-1]} {self.current_year}"

        inc = db.fetch_one(
            f"SELECT COALESCE(SUM(total),0) as total FROM ventas "
            f"WHERE {date_filter} AND estado='completada'", ()
        )
        ingresos_mes = float(inc['total']) if inc else 0.0

        cst = db.fetch_one(
            f"SELECT COALESCE(SUM(dv.cantidad*p.costo),0) as total_costo "
            f"FROM detalle_ventas dv "
            f"JOIN productos p ON dv.producto_id=p.id "
            f"JOIN ventas v    ON dv.venta_id=v.id "
            f"WHERE {date_filter_v} AND v.estado='completada'", ()
        )
        gastos_mes   = float(cst['total_costo']) if cst else 0.0
        ganancia_mes = ingresos_mes - gastos_mes
        margen_mes   = (ganancia_mes / ingresos_mes * 100) if ingresos_mes > 0 else 0.0

        selected_cat = self.category_combo.currentData()

        query = (
            "SELECT c.nombre as categoria, p.nombre as producto, "
            "p.costo as costo_unitario, "
            "SUM(dv.cantidad) as unidades_vendidas, "
            "SUM(dv.subtotal) as ingresos, "
            "SUM(dv.cantidad*p.costo) as gastos "
            "FROM detalle_ventas dv "
            "JOIN productos p  ON dv.producto_id=p.id "
            "JOIN categorias c ON p.categoria_id=c.id "
            "JOIN ventas v     ON dv.venta_id=v.id "
            f"WHERE {date_filter_v} AND v.estado='completada'"
        )
        params = []
        if selected_cat is not None:
            query += " AND c.id=?"
            params.append(selected_cat)
        query += " GROUP BY c.nombre, p.nombre, p.costo ORDER BY c.nombre, ingresos DESC"

        products = db.fetch_all(query, tuple(params))

        # ── Construir filas con totales por categoría ──────────────────
        rows_to_add = []
        current_category = None
        cat_ing = cat_gas = 0.0
        cat_products = []

        for product in products:
            categoria = product['categoria']

            if categoria != current_category:
                # Cerrar categoría anterior
                if current_category is not None:
                    cat_gan = cat_ing - cat_gas
                    cat_mar = (cat_gan / cat_ing * 100) if cat_ing > 0 else 0.0
                    rows_to_add.append({
                        'type':     'category',
                        'label':    f"📁 {current_category}",
                        'ingresos': cat_ing,
                        'gastos':   cat_gas,
                        'ganancia': cat_gan,
                        'margen':   cat_mar,
                    })
                    rows_to_add.extend(cat_products)
                    cat_products = []
                    cat_ing = cat_gas = 0.0

                current_category = categoria

            ingresos_prod = float(product['ingresos'] or 0)
            gastos_prod   = float(product['gastos']   or 0)
            ganancia_prod = ingresos_prod - gastos_prod
            margen_prod   = (ganancia_prod / ingresos_prod * 100) if ingresos_prod > 0 else 0.0

            cat_ing += ingresos_prod
            cat_gas += gastos_prod

            cat_products.append({
                'type':     'product',
                'label':    f"  └─ {product['producto']}",
                'ingresos': ingresos_prod,
                'gastos':   gastos_prod,
                'ganancia': ganancia_prod,
                'margen':   margen_prod,
                'unidades': int(product['unidades_vendidas']),
            })

        # Cerrar última categoría
        if current_category is not None:
            cat_gan = cat_ing - cat_gas
            cat_mar = (cat_gan / cat_ing * 100) if cat_ing > 0 else 0.0
            rows_to_add.append({
                'type':     'category',
                'label':    f"📁 {current_category}",
                'ingresos': cat_ing,
                'gastos':   cat_gas,
                'ganancia': cat_gan,
                'margen':   cat_mar,
            })
            rows_to_add.extend(cat_products)

        # ── Renderizar tabla ───────────────────────────────────────────
        self.analysis_table.setRowCount(1 + len(rows_to_add))

        # Fila 0: resumen del mes
        self._set_table_row(0, label_periodo, ingresos_mes, gastos_mes,
                            ganancia_mes, margen_mes, unidades=None, is_summary=True)

        # Resto de filas
        for i, row_data in enumerate(rows_to_add, 1):
            if row_data['type'] == 'category':
                self._set_table_row(
                    i, row_data['label'],
                    row_data['ingresos'], row_data['gastos'],
                    row_data['ganancia'], row_data['margen'],
                    unidades=None, is_summary=False
                )
            else:
                self._set_table_row(
                    i, row_data['label'],
                    row_data['ingresos'], row_data['gastos'],
                    row_data['ganancia'], row_data['margen'],
                    unidades=row_data['unidades'], is_summary=False
                )
    # ── Callbacks ─────────────────────────────────────────────────────

    def on_month_changed(self):
        self.current_month = self.month_combo.currentData()
        self.day_combo.setCurrentIndex(0)   # resetear día al cambiar mes
        self.costos_card.load_costs(self.current_year, self.current_month)
        self.update_stats()

    def on_year_changed(self):
        self.current_year = self.year_spin.value()
        self.day_combo.setCurrentIndex(0)   # resetear día al cambiar año
        self.costos_card.load_costs(self.current_year, self.current_month)
        self.update_stats()

    def on_category_changed(self):
        self.update_analysis_table()

    def on_day_changed(self):
        self.update_analysis_table()

    def refresh_after_cost_change(self):
        """Llamado por CostosCard o VerCostosDialog tras cambios en gastos."""
        self.costos_card.load_costs(self.current_year, self.current_month)
        self.update_stats()

    # ── Exportar ──────────────────────────────────────────────────────

    def export_excel(self):
        filename = None
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = f"{MESES_ES[self.current_month-1]} {self.current_year}"

            # ── Colores ───────────────────────────────────────────────
            COLOR_HEADER   = "FF6B35"
            COLOR_SUMMARY  = "FFF7ED"
            COLOR_CAT      = "F3F4F6"
            COLOR_GREEN    = "10B981"
            COLOR_RED      = "EF4444"
            COLOR_GRAY     = "6B7280"
            COLOR_WHITE    = "FFFFFF"

            thin = Side(style="thin", color="E5E7EB")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            def header_cell(ws, row, col, value):
                c = ws.cell(row, col, value)
                c.font      = Font(bold=True, color=COLOR_WHITE, size=11)
                c.fill      = PatternFill(start_color=COLOR_HEADER,
                                        end_color=COLOR_HEADER, fill_type="solid")
                c.alignment = Alignment(horizontal="center", vertical="center",
                                        wrap_text=True)
                c.border    = border
                return c

            def data_cell(ws, row, col, value, bold=False, color=None,
                        bg=None, align="right", num_format=None):
                c = ws.cell(row, col, value)
                c.font      = Font(bold=bold, color=color or "1F2937", size=10)
                c.alignment = Alignment(horizontal=align, vertical="center")
                c.border    = border
                if bg:
                    c.fill = PatternFill(start_color=bg, end_color=bg,
                                        fill_type="solid")
                if num_format:
                    c.number_format = num_format
                return c

            # ── Título del reporte ─────────────────────────────────────
            ws.merge_cells("A1:E1")
            title_cell = ws["A1"]
            title_cell.value     = f"Reporte Financiero — {MESES_ES[self.current_month-1]} {self.current_year}"
            title_cell.font      = Font(bold=True, size=14, color=COLOR_HEADER)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            title_cell.fill      = PatternFill(start_color="FFF7ED",
                                            end_color="FFF7ED", fill_type="solid")
            ws.row_dimensions[1].height = 30

            ws.merge_cells("A2:E2")
            gen_cell = ws["A2"]
            gen_cell.value     = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            gen_cell.font      = Font(size=9, color=COLOR_GRAY, italic=True)
            gen_cell.alignment = Alignment(horizontal="right")

            # ── Encabezados ────────────────────────────────────────────
            headers = ["Producto / Categoría", "Ingresos (Bs)",
                    "Gastos (Bs)", "Ganancia Neta (Bs)", "Margen (%)"]
            for col, h in enumerate(headers, 1):
                header_cell(ws, 3, col, h)
            ws.row_dimensions[3].height = 32

            # ── Datos desde la tabla ───────────────────────────────────
            excel_row = 4
            for row in range(self.analysis_table.rowCount()):
                label_item = self.analysis_table.item(row, 0)
                if not label_item:
                    continue

                label = label_item.text()
                is_summary = row == 0
                is_cat     = "📁" in label

                # Leer valores numéricos limpiando "Bs " y "%"
                def get_val(col):
                    item = self.analysis_table.item(row, col)
                    if not item or not item.text():
                        return None
                    txt = item.text().replace("Bs ", "").replace(",", "").replace("%", "").strip()
                    try:
                        return float(txt)
                    except ValueError:
                        return None

                ing = get_val(1)
                gas = get_val(2)
                gan = get_val(3)
                mar = get_val(4)

                bg = ("FFF7ED" if is_summary else
                    "F3F4F6" if is_cat else COLOR_WHITE)

                # Col A — Label
                c = data_cell(ws, excel_row, 1, label,
                            bold=is_summary or is_cat,
                            color="1F2937", bg=bg, align="left")
                if is_summary:
                    c.font = Font(bold=True, size=12, color=COLOR_HEADER)

                # Cols B-E — Números
                if ing is not None:
                    num_fmt = '#,##0.00'
                    ing_col = COLOR_GREEN if not is_cat else "1F2937"
                    gas_col = COLOR_RED   if not is_cat else "1F2937"
                    gan_col = COLOR_GREEN if (gan or 0) >= 0 else COLOR_RED
                    mar_col = COLOR_GREEN if (mar or 0) >= 0 else COLOR_RED

                    data_cell(ws, excel_row, 2, ing, bold=is_summary or is_cat,
                            color=ing_col, bg=bg, num_format=num_fmt)
                    data_cell(ws, excel_row, 3, gas, bold=is_summary or is_cat,
                            color=gas_col, bg=bg, num_format=num_fmt)
                    data_cell(ws, excel_row, 4, gan, bold=is_summary or is_cat,
                            color=gan_col, bg=bg, num_format=num_fmt)
                    pct_cell = data_cell(ws, excel_row, 5, (mar or 0) / 100,
                                        bold=is_summary or is_cat,
                                        color=mar_col, bg=bg,
                                        num_format='0.0%')
                else:
                    for col in range(2, 6):
                        data_cell(ws, excel_row, col, "", bg=bg)

                ws.row_dimensions[excel_row].height = (
                    24 if is_summary else 20 if is_cat else 18)
                excel_row += 1

            # ── Sección resumen de costos ──────────────────────────────
            excel_row += 1
            ws.merge_cells(f"A{excel_row}:E{excel_row}")
            sec = ws.cell(excel_row, 1, "💼 Resumen de Costos del Período")
            sec.font      = Font(bold=True, size=11, color=COLOR_WHITE)
            sec.fill      = PatternFill(start_color="8B5CF6",
                                        end_color="8B5CF6", fill_type="solid")
            sec.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[excel_row].height = 24
            excel_row += 1

            costo_fijo, costo_variable = self.costos_card.get_totals()
            costos_rows = [
                ("Costos Fijos",    costo_fijo,    "3B82F6"),
                ("Costos Variables", costo_variable, "F59E0B"),
                ("Total Costos",    costo_fijo + costo_variable, "EF4444"),
            ]
            for nombre, valor, color in costos_rows:
                data_cell(ws, excel_row, 1, nombre, bold=nombre.startswith("Total"),
                        color="1F2937", align="left")
                data_cell(ws, excel_row, 2, valor,
                        bold=nombre.startswith("Total"),
                        color=color, num_format='#,##0.00')
                for col in range(3, 6):
                    data_cell(ws, excel_row, col, "")
                ws.row_dimensions[excel_row].height = 18
                excel_row += 1

            # ── Anchos de columna ──────────────────────────────────────
            col_widths = [42, 18, 18, 20, 14]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

            # ── Congelar encabezados ───────────────────────────────────
            ws.freeze_panes = "A4"

            output_dir = Path.home() / 'Desktop' / 'La Placita' / 'Reportes Finanzas Excel'
            output_dir.mkdir(parents=True, exist_ok=True)
            if not filename:
                filename = str(output_dir / f"finanzas_{self.current_year}_{self.current_month:02d}.xlsx")
            else:
                filename = str(output_dir / Path(filename).name)

            wb.save(filename)
            QMessageBox.information(self, "Éxito", f"Excel generado:\n{filename}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al generar Excel:\n{str(e)}")

    def export_pdf(self):
        filename = None
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import cm
            from reportlab.lib import colors
            from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                            Paragraph, Spacer, HRFlowable)
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
            

            # ── Archivo de salida ──────────────────────────────────────

            output_dir = Path.home() / 'Desktop' / 'La Placita' / 'Reportes Finanzas PDF'
            output_dir.mkdir(parents=True, exist_ok=True)
            if not filename:
                filename = str(output_dir / f"finanzas_{self.current_year}_{self.current_month:02d}.pdf")
            else:
                filename = str(output_dir / Path(filename).name)

            doc = SimpleDocTemplate(
                filename, pagesize=A4,
                rightMargin=2*cm, leftMargin=2*cm,
                topMargin=2*cm, bottomMargin=2*cm
            )

            # ── Estilos ────────────────────────────────────────────────
            C_ORANGE  = colors.HexColor("#FF6B35")
            C_GREEN   = colors.HexColor("#10B981")
            C_RED     = colors.HexColor("#EF4444")
            C_DARK    = colors.HexColor("#1F2937")
            C_GRAY    = colors.HexColor("#6B7280")
            C_LGRAY   = colors.HexColor("#F3F4F6")
            C_WHITE   = colors.white
            C_PURPLE  = colors.HexColor("#8B5CF6")

            styles = getSampleStyleSheet()

            def style(name, parent="Normal", **kw):
                return ParagraphStyle(name, parent=styles[parent], **kw)

            s_title    = style("Title2",    fontSize=18, textColor=C_ORANGE,
                            fontName="Helvetica-Bold", spaceAfter=4, alignment=TA_LEFT)
            s_subtitle = style("Sub2",      fontSize=10, textColor=C_GRAY,
                            spaceAfter=16, alignment=TA_LEFT)
            s_section  = style("Section",   fontSize=10, textColor=C_DARK,
                            fontName="Helvetica-Bold", spaceBefore=14, spaceAfter=6)
            s_footer   = style("Footer",    fontSize=8,  textColor=C_GRAY,
                            alignment=TA_CENTER, spaceBefore=12)

            elements = []

            # ── Encabezado ─────────────────────────────────────────────
            elements.append(Paragraph(
                f"Reporte Financiero — {MESES_ES[self.current_month-1]} {self.current_year}",
                s_title))
            elements.append(Paragraph(
                f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}  ·  "
                f"La Placita Cafetería",
                s_subtitle))
            elements.append(HRFlowable(width="100%", thickness=1,
                                    color=C_ORANGE, spaceAfter=12))

            # ── Resumen ejecutivo ──────────────────────────────────────
            month_str = f"{self.current_year}-{self.current_month:02d}"
            inc = db.fetch_one(
                "SELECT COALESCE(SUM(total),0) as total FROM ventas "
                "WHERE strftime('%Y-%m',fecha_venta)=? AND estado='completada'",
                (month_str,)
            )
            ingresos = float(inc['total']) if inc else 0.0

            cst = db.fetch_one(
                "SELECT COALESCE(SUM(dv.cantidad*p.costo),0) as total_costo "
                "FROM detalle_ventas dv "
                "JOIN productos p ON dv.producto_id=p.id "
                "JOIN ventas v    ON dv.venta_id=v.id "
                "WHERE strftime('%Y-%m',v.fecha_venta)=? AND v.estado='completada'",
                (month_str,)
            )
            gastos_prod = float(cst['total_costo']) if cst else 0.0

            costo_fijo, costo_variable = self.costos_card.get_totals()
            gastos_total = gastos_prod + costo_fijo + costo_variable
            ganancia     = ingresos - gastos_total
            margen       = (ganancia / ingresos * 100) if ingresos > 0 else 0.0

            elements.append(Paragraph("Resumen del Período", s_section))

            summary_data = [
                ["Concepto", "Monto (Bs)"],
                ["Ingresos Totales",          f"Bs {ingresos:,.2f}"],
                ["Costos de Productos",        f"Bs {gastos_prod:,.2f}"],
                ["Costos Fijos",               f"Bs {costo_fijo:,.2f}"],
                ["Costos Variables",           f"Bs {costo_variable:,.2f}"],
                ["Total Gastos",               f"Bs {gastos_total:,.2f}"],
                ["Ganancia Neta",              f"Bs {ganancia:,.2f}"],
                ["Margen de Ganancia",         f"{margen:.1f}%"],
            ]

            gan_color = C_GREEN if ganancia >= 0 else C_RED
            mar_color = C_GREEN if margen   >= 0 else C_RED

            summary_table = Table(summary_data, colWidths=[11*cm, 5*cm])
            summary_table.setStyle(TableStyle([
                # Encabezado
                ("BACKGROUND",  (0,0), (-1,0), C_DARK),
                ("TEXTCOLOR",   (0,0), (-1,0), C_WHITE),
                ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",    (0,0), (-1,0), 9),
                ("ALIGN",       (1,0), (1,0),  "RIGHT"),
                # Cuerpo
                ("FONTSIZE",    (0,1), (-1,-1), 9),
                ("FONTNAME",    (0,1), (0,-1),  "Helvetica"),
                ("ALIGN",       (1,1), (1,-1),  "RIGHT"),
                ("ROWBACKGROUNDS", (0,1), (-1,-1), [C_WHITE, C_LGRAY]),
                ("TOPPADDING",  (0,0), (-1,-1), 6),
                ("BOTTOMPADDING",(0,0),(-1,-1), 6),
                ("LEFTPADDING", (0,0), (-1,-1), 8),
                ("GRID",        (0,0), (-1,-1), 0.5, colors.HexColor("#E5E7EB")),
                # Fila Total Gastos — negrita
                ("FONTNAME",    (0,5), (-1,5), "Helvetica-Bold"),
                ("TEXTCOLOR",   (1,5), (1,5),  C_RED),
                # Fila Ganancia — color dinámico
                ("FONTNAME",    (0,6), (-1,6), "Helvetica-Bold"),
                ("TEXTCOLOR",   (1,6), (1,6),  gan_color),
                # Fila Margen
                ("FONTNAME",    (0,7), (-1,7), "Helvetica-Bold"),
                ("TEXTCOLOR",   (1,7), (1,7),  mar_color),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 0.5*cm))

            # ── Desglose por producto ──────────────────────────────────
            elements.append(HRFlowable(width="100%", thickness=0.5,
                                    color=C_GRAY, spaceAfter=6))
            elements.append(Paragraph("Desglose por Producto", s_section))

            detail_data = [["Producto", "Unidades", "Ingresos", "Gastos",
                            "Ganancia", "Margen"]]

            for row in range(self.analysis_table.rowCount()):
                label_item = self.analysis_table.item(row, 0)
                if not label_item:
                    continue
                label = label_item.text()
                is_summary = row == 0
                if is_summary:
                    continue  # ya está en el resumen

                def get_txt(col):
                    item = self.analysis_table.item(row, col)
                    return item.text() if item else ""

                if "📁" in label:
                    # Fila de categoría — resaltada
                    clean = label.replace("📁", "").strip()
                    detail_data.append([clean, "", get_txt(1), get_txt(2),
                                        get_txt(3), get_txt(4)])
                else:
                    # Fila de producto — extraer unidades del label
                    clean = label.replace("└─", "").strip()
                    # Las unidades están al final del label entre paréntesis
                    import re
                    match = re.search(r'\((\d+) unidades\)', clean)
                    unidades = match.group(1) if match else "—"
                    nombre   = re.sub(r'\s*\(\d+ unidades\)', '', clean).strip()
                    detail_data.append([nombre, unidades, get_txt(1), get_txt(2),
                                        get_txt(3), get_txt(4)])

            col_widths_detail = [6.5*cm, 1.8*cm, 2.8*cm, 2.8*cm, 2.8*cm, 2*cm]
            detail_table = Table(detail_data, colWidths=col_widths_detail,
                                repeatRows=1)

            # Estilo base
            ts = TableStyle([
                ("BACKGROUND",   (0,0), (-1,0), C_DARK),
                ("TEXTCOLOR",    (0,0), (-1,0), C_WHITE),
                ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",     (0,0), (-1,-1), 8),
                ("ALIGN",        (1,0), (-1,-1), "RIGHT"),
                ("ALIGN",        (0,0), (0,-1),  "LEFT"),
                ("ROWBACKGROUNDS",(0,1),(-1,-1), [C_WHITE, C_LGRAY]),
                ("TOPPADDING",   (0,0), (-1,-1), 5),
                ("BOTTOMPADDING",(0,0), (-1,-1), 5),
                ("LEFTPADDING",  (0,0), (-1,-1), 6),
                ("GRID",         (0,0), (-1,-1), 0.3, colors.HexColor("#E5E7EB")),
            ])

            # Resaltar filas de categoría
            for i, row_data in enumerate(detail_data[1:], 1):
                label_val = str(row_data[0])
                # Detectar si era categoría (sin unidades y sin └─ ya limpiamos)
                item = self.analysis_table.item(i, 0)  # fila original +1 por summary
                if item and "📁" in item.text():
                    ts.add("FONTNAME",   (0,i), (-1,i), "Helvetica-Bold")
                    ts.add("BACKGROUND", (0,i), (-1,i), colors.HexColor("#E5E7EB"))
                    ts.add("TEXTCOLOR",  (0,i), (-1,i), C_DARK)

            detail_table.setStyle(ts)
            elements.append(detail_table)

            # ── Costos registrados ─────────────────────────────────────
            elements.append(Spacer(1, 0.4*cm))
            elements.append(HRFlowable(width="100%", thickness=0.5,
                                    color=C_GRAY, spaceAfter=6))
            elements.append(Paragraph("Costos Registrados del Período", s_section))

            costos_rows_db = db.fetch_all(
                "SELECT tipo, concepto, categoria, monto, fecha_gasto "
                "FROM gastos WHERE strftime('%Y-%m', fecha_gasto)=? "
                "ORDER BY tipo, fecha_gasto",
                (month_str,)
            )

            if costos_rows_db:
                costos_data = [["Tipo", "Concepto", "Categoría", "Fecha", "Monto"]]
                for r in costos_rows_db:
                    tipo = "Fijo" if r['tipo'] == 'fijo' else "Variable"
                    try:
                        fecha = datetime.strptime(
                            r['fecha_gasto'], "%Y-%m-%d").strftime("%d/%m/%Y")
                    except Exception:
                        fecha = r['fecha_gasto']
                    costos_data.append([
                        tipo,
                        r['concepto'],
                        r['categoria'] or "—",
                        fecha,
                        f"Bs {float(r['monto']):,.2f}",
                    ])

                costos_table = Table(
                    costos_data,
                    colWidths=[2.2*cm, 6*cm, 3.5*cm, 2.5*cm, 2.5*cm],
                    repeatRows=1
                )
                costos_table.setStyle(TableStyle([
                    ("BACKGROUND",    (0,0), (-1,0), colors.HexColor("#374151")),
                    ("TEXTCOLOR",     (0,0), (-1,0), C_WHITE),
                    ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
                    ("FONTSIZE",      (0,0), (-1,-1), 8),
                    ("ALIGN",         (4,1), (4,-1),  "RIGHT"),
                    ("ALIGN",         (0,0), (3,-1),  "LEFT"),
                    ("ROWBACKGROUNDS",(0,1), (-1,-1), [C_WHITE, C_LGRAY]),
                    ("TOPPADDING",    (0,0), (-1,-1), 5),
                    ("BOTTOMPADDING", (0,0), (-1,-1), 5),
                    ("LEFTPADDING",   (0,0), (-1,-1), 6),
                    ("GRID",          (0,0), (-1,-1), 0.3,
                    colors.HexColor("#E5E7EB")),
                ]))
                elements.append(costos_table)
            else:
                elements.append(Paragraph(
                    "No hay costos registrados para este período.",
                    style("NoData", fontSize=9, textColor=C_GRAY)))

            # ── Pie de página ──────────────────────────────────────────
            elements.append(Spacer(1, 0.8*cm))
            elements.append(HRFlowable(width="100%", thickness=0.5, color=C_LGRAY))
            elements.append(Paragraph(
                "La Placita Cafetería  ·  Reporte generado automáticamente  ·  "
                f"{datetime.now().strftime('%d/%m/%Y')}",
                s_footer))

            doc.build(elements)
            QMessageBox.information(self, "Éxito", f"PDF generado:\n{filename}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al generar PDF:\n{str(e)}")